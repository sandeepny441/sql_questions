from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable
import re

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parent
SOURCES_DIR = ROOT / "tmp" / "sources"
OUTPUT_PATH = ROOT / "MBS_pooling_dictionary.xlsx"


@dataclass(frozen=True)
class Source:
    title: str
    url: str
    local_path: Path | None = None


SOURCES = {
    "fannie_glossary_a": Source(
        title="Fannie Mae Glossary: A",
        url="https://selling-guide.fanniemae.com/glossary/a",
        local_path=SOURCES_DIR / "fn_glossary_a.html",
    ),
    "fannie_glossary_c": Source(
        title="Fannie Mae Glossary: C",
        url="https://selling-guide.fanniemae.com/glossary/c",
        local_path=SOURCES_DIR / "fn_glossary_c.html",
    ),
    "fannie_glossary_d": Source(
        title="Fannie Mae Glossary: D",
        url="https://selling-guide.fanniemae.com/glossary/d",
        local_path=SOURCES_DIR / "fn_glossary_d.html",
    ),
    "fannie_glossary_g": Source(
        title="Fannie Mae Glossary: G",
        url="https://selling-guide.fanniemae.com/glossary/g",
        local_path=SOURCES_DIR / "fn_glossary_g.html",
    ),
    "fannie_glossary_i": Source(
        title="Fannie Mae Glossary: I",
        url="https://selling-guide.fanniemae.com/glossary/i",
        local_path=SOURCES_DIR / "fn_glossary_i.html",
    ),
    "fannie_glossary_m": Source(
        title="Fannie Mae Glossary: M",
        url="https://selling-guide.fanniemae.com/glossary/m",
        local_path=SOURCES_DIR / "fn_glossary_m.html",
    ),
    "fannie_glossary_p": Source(
        title="Fannie Mae Glossary: P",
        url="https://selling-guide.fanniemae.com/glossary/p",
        local_path=SOURCES_DIR / "fn_glossary_p.html",
    ),
    "fannie_glossary_r": Source(
        title="Fannie Mae Glossary: R",
        url="https://selling-guide.fanniemae.com/glossary/r",
        local_path=SOURCES_DIR / "fn_glossary_r.html",
    ),
    "fannie_glossary_s": Source(
        title="Fannie Mae Glossary: S",
        url="https://selling-guide.fanniemae.com/glossary/s",
        local_path=SOURCES_DIR / "fn_glossary_s.html",
    ),
    "fannie_glossary_s_servicing": Source(
        title="Fannie Mae Servicing Glossary: S",
        url="https://servicing-guide.fanniemae.com/svc/f-3-19/acronyms-and-glossary-defined-terms-s",
        local_path=SOURCES_DIR / "fn_glossary_s_servicing.html",
    ),
    "fannie_glossary_t": Source(
        title="Fannie Mae Glossary: T",
        url="https://servicing-guide.fanniemae.com/svc/f-3-20/acronyms-and-glossary-defined-terms-t",
        local_path=SOURCES_DIR / "fn_glossary_t.html",
    ),
    "fannie_glossary_u": Source(
        title="Fannie Mae Glossary: U",
        url="https://servicing-guide.fanniemae.com/svc/f-3-21/acronyms-and-glossary-defined-terms-u",
        local_path=SOURCES_DIR / "fn_glossary_u.html",
    ),
    "fannie_glossary_w": Source(
        title="Fannie Mae Glossary: W",
        url="https://selling-guide.fanniemae.com/glossary/w",
        local_path=SOURCES_DIR / "fn_glossary_w.html",
    ),
    "fannie_c3_1_01": Source(
        title="General Information About Fannie Mae's MBS Program",
        url="https://selling-guide.fanniemae.com/sel/c3-1-01/general-information-about-fannie-maes-mbs-program",
        local_path=SOURCES_DIR / "fn_c3_1_01.html",
    ),
    "fannie_good_delivery": Source(
        title="Making Good Delivery",
        url="https://selling-guide.fanniemae.com/sel/c3-7-03/making-good-delivery",
        local_path=SOURCES_DIR / "fn_good_delivery.html",
    ),
    "fannie_remittance": Source(
        title="MBS Remittance Type and Selecting a Remittance Cycle",
        url="https://selling-guide.fanniemae.com/sel/c3-2-03/mbs-remittance-type-and-selecting-remittance-cycle",
        local_path=SOURCES_DIR / "fn_remittance.html",
    ),
    "fannie_buyup_buydown": Source(
        title="Buying Up and Buying Down the Guaranty Fee for MBS",
        url="https://selling-guide.fanniemae.com/sel/c3-3-03/buying-and-buying-down-guaranty-fee-mbs",
        local_path=SOURCES_DIR / "fn_buyup_buydown.html",
    ),
    "fannie_disclosure_guide": Source(
        title="Fannie Mae Single-Family Disclosure Guide",
        url="https://capitalmarkets.fanniemae.com/resources/file/mbs/pdf/mbsglossary.pdf",
    ),
    "user_pdf": Source(
        title="MBS Pool Transactions Overview updated Jan 2024.pdf",
        url=str((ROOT / "MBS Pool Transactions Overview updated Jan 2024.pdf").resolve()),
    ),
}


def normalize_term(text: str) -> str:
    return re.sub(r"\s+", " ", text.strip().lower())


def clean_text(text: str) -> str:
    text = text.replace("\xa0", " ")
    text = text.replace("’", "'")
    text = text.replace("“", '"').replace("”", '"')
    text = text.replace("–", "-").replace("—", "-")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def parse_glossary_rows(source: Source) -> dict[str, dict[str, str]]:
    if not source.local_path or not source.local_path.exists():
        return {}

    soup = BeautifulSoup(source.local_path.read_text(errors="ignore"), "lxml")
    rows: dict[str, dict[str, str]] = {}
    for row in soup.select(".views-row"):
        title_node = row.select_one(".views-field-title .field-content")
        definition_node = row.select_one(".views-field-field-definition .field-content")
        if not title_node or not definition_node:
            continue
        term = clean_text(title_node.get_text(" ", strip=True))
        definition = clean_text(definition_node.get_text(" ", strip=True))
        if not term or not definition:
            continue
        rows[normalize_term(term)] = {
            "term": term,
            "definition": definition,
            "source_title": source.title,
            "source_url": source.url,
        }
    return rows


def build_official_lookup() -> dict[str, dict[str, str]]:
    lookup: dict[str, dict[str, str]] = {}
    official_keys = {key for key in SOURCES if key.startswith("fannie_glossary_")}
    official_keys.add("fannie_glossary_s_servicing")
    for key in sorted(official_keys):
        source = SOURCES[key]
        lookup.update(parse_glossary_rows(source))
    return lookup


OFFICIAL_LOOKUP = build_official_lookup()


def extract_acronym(term: str) -> str:
    paren_match = re.search(r"\(([A-Z0-9&/\- ]{2,})\)", term)
    if paren_match:
        return paren_match.group(1).strip()
    if term.isupper() and 2 <= len(term) <= 8:
        return term
    return ""


def infer_category(term: str) -> str:
    t = normalize_term(term)
    if any(k in t for k in ("remittance", "amortization", "maturity", "wac", "wala", "wam", "factor")):
        return "Rates, Cash Flow & Analytics"
    if any(k in t for k in ("guaranty", "servicing", "coupon", "accrual", "margin", "basis point", "upb", "face", "wac", "wala", "wam", "yield", "principal", "interest")):
        return "Rates, Cash Flow & Analytics"
    if any(k in t for k in ("trade", "settlement", "delivery", "good delivery", "pair-off", "fail", "tba", "specified pool", "variance", "allocation", "lot")):
        return "Trading & Settlement"
    if any(k in t for k in ("pool", "securitization", "swap", "commitment", "prefix", "factor")):
        return "Pooling Structure & Workflow"
    if any(k in t for k in ("security", "mbs", "umbs", "megas", "smbs", "cusip", "book-entry")):
        return "Security Structure"
    if any(k in t for k in ("loan", "mortgage", "borrower", "lender", "seller", "servicer", "investor", "underwriting", "eligibility", "delinquency", "seasoned", "nonstandard")):
        return "Collateral, Parties & Eligibility"
    if any(k in t for k in ("document", "desk", "sifma", "federal reserve", "notification")):
        return "Operations & Market Infrastructure"
    return "General"


def make_row(
    term: str,
    *,
    definition: str | None = None,
    source_key: str | None = None,
    lookup_term: str | None = None,
    notes: str = "",
    category: str | None = None,
) -> dict[str, str]:
    resolved_source = SOURCES[source_key] if source_key else None
    official = OFFICIAL_LOOKUP.get(normalize_term(lookup_term or term))

    if definition is None and official:
        definition = official["definition"]
    if definition is None:
        raise ValueError(f"Missing definition for term: {term}")

    source_title = ""
    source_url = ""
    source_type = "Operational summary"
    if official:
        source_title = official["source_title"]
        source_url = official["source_url"]
        source_type = "Official glossary definition"
    elif resolved_source:
        source_title = resolved_source.title
        source_url = resolved_source.url

    return {
        "Term": term,
        "Acronym": extract_acronym(term),
        "Category": category or infer_category(term),
        "Definition": clean_text(definition),
        "Notes": clean_text(notes),
        "Source Type": source_type,
        "Source Title": source_title,
        "Source URL": source_url,
    }


def official_rows() -> list[dict[str, str]]:
    terms = [
        "accrual rate",
        "actual/actual remittance type",
        "adjustable-rate mortgage (ARM)",
        "ARM",
        "ARM Flex",
        "ARM Flex Plus",
        "cash back pair-off",
        "CUSIP",
        "CUSIP number",
        "delivery versus payment settlement",
        "Document Certification",
        "document custodian",
        "good delivery",
        "guaranty fee",
        "guaranty fee buydown",
        "guaranty fee buyup",
        "index",
        "issue date",
        "issue date principal balance",
        "market-rate option",
        "master servicer",
        "maximum pool accrual rate",
        "maximum weighted-average pool accrual rate",
        "MBS",
        "MBS Express pool",
        "MBS Express remittance cycle",
        "MBS margin",
        "MBS mortgage",
        "MBS pool",
        "MBS pool delivery",
        "Megas",
        "mortgage-backed security (MBS)",
        "mortgage loan",
        "mortgage note",
        "multiple pool",
        "pair-off",
        "pass-through rate",
        "pool",
        "pool accrual rate",
        "pool issue date",
        "pool purchase contract",
        "pool purchase transaction",
        "portfolio mortgage",
        "principal distribution amount",
        "rapid payment method (RPM)",
        "remittance cycle",
        "repurchase date",
        "scheduled/scheduled remittance type",
        "security",
        "security balance",
        "servicing compensation",
        "servicing fee",
        "servicing spread",
        "settlement date",
        "SMBS",
        "standard pricing option",
        "standard remittance cycle",
        "stated-structure pooling",
        "TBA",
        "take-out option",
        "Uniform Mortgage-Backed Security (UMBS)",
        "UPB",
        "pool transaction amount",
        "weighted-average pool accrual rate",
        "weighted-average structure pooling",
        "whole loan delivery",
    ]
    return [make_row(term) for term in terms]


def manual_rows() -> list[dict[str, str]]:
    rows = [
        make_row(
            "Agency MBS",
            definition="A mortgage-backed security issued or guaranteed by a housing agency or government-sponsored enterprise, typically Fannie Mae, Freddie Mac, or Ginnie Mae.",
            source_key="fannie_c3_1_01",
        ),
        make_row(
            "Allocation",
            definition="The process of matching specific pools and balances to a particular trade or set of trades for settlement.",
            source_key="user_pdf",
            category="Trading & Settlement",
        ),
        make_row(
            "Allocation sheet",
            definition="The worksheet used to show which pools, amounts, prices, and trade numbers are being assigned to a delivery.",
            source_key="user_pdf",
            category="Trading & Settlement",
        ),
        make_row(
            "Amortization schedule",
            definition="The month-by-month schedule showing how a mortgage pays down principal and interest over time.",
            source_key="fannie_c3_1_01",
        ),
        make_row(
            "Basis point",
            definition="One one-hundredth of one percent, or 0.01%; MBS fees, coupons, and variances are often discussed in basis points.",
            source_key="fannie_buyup_buydown",
        ),
        make_row(
            "Book-entry delivery",
            definition="Electronic delivery of securities through an account system rather than by moving paper certificates.",
            source_key="fannie_glossary_c",
        ),
        make_row(
            "Borrower",
            definition="The homeowner or property owner who owes payments on the mortgage loan.",
            source_key="fannie_c3_1_01",
        ),
        make_row(
            "Existing issue",
            definition="A trade in already-issued MBS, often settled by delivery-versus-payment rather than as a new issuance.",
            source_key="fannie_glossary_d",
            category="Trading & Settlement",
        ),
        make_row(
            "FHA",
            definition="The Federal Housing Administration, whose insured loans can back certain agency mortgage securities.",
            source_key="fannie_c3_1_01",
            category="Operations & Market Infrastructure",
        ),
        make_row(
            "Capital Markets Pricing and Sales Desk",
            definition="The Fannie Mae trading and execution desk that prices trades, receives trade-related communications, and manages delivery operations.",
            source_key="user_pdf",
            category="Operations & Market Infrastructure",
        ),
        make_row(
            "Collateral",
            definition="The underlying mortgage loans that back an MBS or any related pooling structure.",
            source_key="fannie_c3_1_01",
        ),
        make_row(
            "Commingling",
            definition="Pooling loans with certain different characteristics together when the program rules specifically allow it.",
            source_key="user_pdf",
        ),
        make_row(
            "Commitment",
            definition="A contractual obligation under which a lender agrees to deliver loans or securities at agreed pricing and terms within a stated time window.",
            source_key="fannie_c3_1_01",
        ),
        make_row(
            "Conventional loan",
            definition="A mortgage that is not insured or guaranteed by a federal agency such as FHA or VA.",
            source_key="fannie_c3_1_01",
        ),
        make_row(
            "Coupon",
            definition="The interest rate associated with the MBS security from the investor's perspective; in agency pass-throughs this is closely tied to the pass-through rate.",
            source_key="fannie_c3_1_01",
        ),
        make_row(
            "Current face",
            definition="The current outstanding principal balance of a security or delivered pool position after principal paydowns.",
            source_key="user_pdf",
            category="Rates, Cash Flow & Analytics",
        ),
        make_row(
            "Current face variance",
            definition="The difference between the current balance actually delivered and the target trade amount, measured against allowed tolerance.",
            source_key="fannie_good_delivery",
            category="Trading & Settlement",
        ),
        make_row(
            "Curtailment",
            definition="A partial unscheduled principal payment made before the mortgage is fully paid off.",
            source_key="fannie_remittance",
            category="Rates, Cash Flow & Analytics",
        ),
        make_row(
            "Delinquency",
            definition="A status in which the borrower has fallen behind on scheduled mortgage payments.",
            source_key="fannie_c3_1_01",
        ),
        make_row(
            "Delivery variance",
            definition="The allowed plus-or-minus difference between the committed trade amount and the amount actually delivered.",
            source_key="fannie_good_delivery",
            category="Trading & Settlement",
        ),
        make_row(
            "Desk",
            definition="Trading-desk shorthand for the capital markets desk that executes and settles MBS trades.",
            source_key="fannie_good_delivery",
            category="Operations & Market Infrastructure",
        ),
        make_row(
            "Fail",
            definition="A settlement breakdown in which the agreed security is not delivered correctly or on time.",
            source_key="fannie_good_delivery",
            category="Trading & Settlement",
        ),
        make_row(
            "Fail charge",
            definition="A cost or penalty that can arise when a trade does not settle on time or does not satisfy delivery requirements.",
            source_key="user_pdf",
            category="Trading & Settlement",
        ),
        make_row(
            "Fannie Mae",
            definition="The government-sponsored enterprise that purchases eligible mortgages, securitizes them into MBS, and guarantees timely principal and interest on its securities.",
            source_key="fannie_c3_1_01",
            category="Operations & Market Infrastructure",
        ),
        make_row(
            "Freddie Mac",
            definition="The other large U.S. housing-finance GSE whose single-family agency MBS compete and co-trade with Fannie Mae securities in the TBA market.",
            source_key="fannie_glossary_u",
            category="Operations & Market Infrastructure",
        ),
        make_row(
            "Fannie Majors",
            definition="Fannie Mae's multiple-lender pooling structure, where several lenders contribute collateral to a common security.",
            source_key="user_pdf",
        ),
        make_row(
            "Federal Reserve book-entry system",
            definition="The electronic account system used to hold and transfer many agency MBS positions without paper certificates.",
            source_key="fannie_glossary_s",
            category="Operations & Market Infrastructure",
        ),
        make_row(
            "Final maturity",
            definition="The month in which the mortgage or security is scheduled to pay down to zero under its contractual terms.",
            source_key="user_pdf",
            category="Rates, Cash Flow & Analytics",
        ),
        make_row(
            "First-lien mortgage",
            definition="A mortgage with the senior claim on the property collateral; this is the standard lien position for loans pooled into agency single-family MBS.",
            source_key="fannie_c3_1_01",
        ),
        make_row(
            "Fixed-rate mortgage (FRM)",
            definition="A mortgage whose note rate does not reset during its term.",
            source_key="fannie_c3_1_01",
        ),
        make_row(
            "Good delivery millions",
            definition="An operational way of thinking about delivery in million-dollar lots, where each lot has to satisfy pool-count and variance rules.",
            source_key="user_pdf",
            category="Trading & Settlement",
        ),
        make_row(
            "Government loan",
            definition="A mortgage insured or guaranteed by a government agency, such as FHA, VA, or RHS/RD.",
            source_key="fannie_c3_1_01",
        ),
        make_row(
            "Ginnie Mae",
            definition="The government agency whose MBS are backed by federally insured or guaranteed loans and which appears in TBA delivery conventions alongside Fannie and Freddie securities.",
            source_key="user_pdf",
            category="Operations & Market Infrastructure",
        ),
        make_row(
            "Investor",
            definition="The person or institution that buys the MBS and receives the cash flow passed through from the underlying pool.",
            source_key="fannie_c3_1_01",
        ),
        make_row(
            "Loan-to-security swap",
            definition="A transaction in which the lender delivers mortgages and receives mortgage-backed securities rather than just a cash purchase price.",
            source_key="fannie_c3_1_01",
        ),
        make_row(
            "Lot amount",
            definition="The target trade amount assigned to a delivery lot, often expressed in one-million-dollar increments for good delivery testing.",
            source_key="user_pdf",
            category="Trading & Settlement",
        ),
        make_row(
            "Mandatory commitment",
            definition="A commitment structure under which the seller is obligated to deliver the agreed loans or securities rather than having an option to walk away.",
            source_key="fannie_c3_1_01",
        ),
        make_row(
            "MBS commitment",
            definition="The governing contract under which a lender delivers collateral for securitization or related MBS execution.",
            source_key="fannie_glossary_p",
            category="Pooling Structure & Workflow",
        ),
        make_row(
            "Minimum delivery amount",
            definition="The smallest balance that still satisfies the applicable delivery rules for the trade and settlement method.",
            source_key="fannie_good_delivery",
            category="Trading & Settlement",
        ),
        make_row(
            "Minimum pool balance",
            definition="The smallest principal balance permitted for a pool or pool submission under the applicable pooling or delivery rules.",
            source_key="user_pdf",
        ),
        make_row(
            "Mortgage loan",
            definition="The individual home loan that serves as the basic unit of collateral in a pool.",
            source_key="fannie_glossary_m",
        ),
        make_row(
            "Mortgage note",
            definition="The legal instrument that records the borrower's obligation to repay the mortgage debt.",
            source_key="fannie_c3_1_01",
        ),
        make_row(
            "Multiple pool deliveries",
            definition="A delivery setup in which more than one pool is used to satisfy a trade, subject to maximum pool-count rules.",
            source_key="user_pdf",
            category="Trading & Settlement",
        ),
        make_row(
            "Nonstandard loan",
            definition="A loan with characteristics that make it subject to special pooling or TBA-delivery limits under agency rules.",
            source_key="fannie_good_delivery",
        ),
        make_row(
            "Original face",
            definition="The principal balance of the security at issuance, before any later paydowns.",
            source_key="user_pdf",
            category="Rates, Cash Flow & Analytics",
        ),
        make_row(
            "Originator",
            definition="The lender or mortgage company that originally makes the loan to the borrower.",
            source_key="fannie_c3_1_01",
        ),
        make_row(
            "P&I",
            definition="Short for principal and interest, the core payment components collected on a mortgage.",
            source_key="fannie_remittance",
            category="Rates, Cash Flow & Analytics",
        ),
        make_row(
            "Pool data and documents",
            definition="The loan-level, pool-level, and certification materials required to support a clean MBS delivery.",
            source_key="fannie_good_delivery",
            category="Operations & Market Infrastructure",
        ),
        make_row(
            "Pool factor",
            definition="The current security balance divided by the original face amount, showing what share of the original balance remains outstanding.",
            source_key="fannie_disclosure_guide",
            category="Rates, Cash Flow & Analytics",
        ),
        make_row(
            "Pool notification",
            definition="The communication sent before settlement identifying which pools will be delivered on a trade.",
            source_key="user_pdf",
            category="Trading & Settlement",
        ),
        make_row(
            "Pool number",
            definition="The unique identifier assigned to a specific mortgage pool or security.",
            source_key="user_pdf",
            category="Pooling Structure & Workflow",
        ),
        make_row(
            "Pool prefix",
            definition="The code embedded in a pool identifier that signals the general product type or security class.",
            source_key="fannie_glossary_u",
            category="Pooling Structure & Workflow",
        ),
        make_row(
            "Prepayment",
            definition="Any principal returned earlier than scheduled, whether from refinance, home sale, curtailment, or payoff.",
            source_key="fannie_remittance",
            category="Rates, Cash Flow & Analytics",
        ),
        make_row(
            "Pass-through certificate",
            definition="Another name for a basic agency MBS because principal and interest from the underlying loans are passed through to investors.",
            source_key="fannie_c3_1_01",
            category="Security Structure",
        ),
        make_row(
            "Seasoned loan",
            definition="A loan that has already been outstanding for some period of time before being pooled or delivered.",
            source_key="fannie_c3_1_01",
        ),
        make_row(
            "Securitization",
            definition="The process of turning mortgage loans into a tradable security backed by those loans.",
            source_key="fannie_c3_1_01",
        ),
        make_row(
            "Seller",
            definition="The institution that delivers the mortgage loans or MBS into the transaction with Fannie Mae or the market counterparty.",
            source_key="fannie_c3_1_01",
        ),
        make_row(
            "Servicer",
            definition="The party that collects payments, manages escrow and delinquency work, and remits funds according to the servicing and remittance rules.",
            source_key="fannie_remittance",
        ),
        make_row(
            "Settlement",
            definition="The point in the trade lifecycle when securities are delivered and funds are exchanged.",
            source_key="fannie_glossary_s",
            category="Trading & Settlement",
        ),
        make_row(
            "Single-lender pool",
            definition="A pool backed by loans delivered by one lender rather than by multiple contributors.",
            source_key="fannie_good_delivery",
        ),
        make_row(
            "SIFMA",
            definition="The Securities Industry and Financial Markets Association, whose market conventions help define TBA good delivery.",
            source_key="fannie_good_delivery",
            category="Operations & Market Infrastructure",
        ),
        make_row(
            "Specified pool",
            definition="A trade or security identified by exact pool number and characteristics rather than left open for later TBA allocation.",
            source_key="fannie_good_delivery",
            category="Trading & Settlement",
        ),
        make_row(
            "Swap",
            definition="Market shorthand for exchanging mortgage collateral for an MBS issuance.",
            source_key="user_pdf",
            category="Pooling Structure & Workflow",
        ),
        make_row(
            "To-be-announced market",
            definition="The forward market in which agency MBS are traded on standard parameters before the exact pool is specified.",
            source_key="fannie_glossary_u",
            category="Trading & Settlement",
            lookup_term="Uniform Mortgage-Backed Security (UMBS)",
        ),
        make_row(
            "Trade amount",
            definition="The principal amount the seller has agreed to deliver against the trade.",
            source_key="fannie_good_delivery",
            category="Trading & Settlement",
        ),
        make_row(
            "Trade assignment",
            definition="The process of assigning all or part of a trade obligation to another party before settlement.",
            source_key="user_pdf",
            category="Trading & Settlement",
        ),
        make_row(
            "Trade date",
            definition="The day the two sides agree to the terms of the MBS transaction.",
            source_key="fannie_good_delivery",
            category="Trading & Settlement",
        ),
        make_row(
            "Underwriting",
            definition="The lender's credit and collateral review process used to determine whether a mortgage loan should be originated or purchased.",
            source_key="fannie_c3_1_01",
        ),
        make_row(
            "Unpaid principal balance (UPB)",
            definition="The remaining principal amount still outstanding on a mortgage loan or security.",
            source_key="fannie_glossary_u",
            category="Rates, Cash Flow & Analytics",
        ),
        make_row(
            "VA",
            definition="The U.S. Department of Veterans Affairs, whose guaranteed loans can be pooled into agency mortgage securities.",
            source_key="fannie_c3_1_01",
            category="Operations & Market Infrastructure",
        ),
        make_row(
            "Variance",
            definition="The tolerance between the committed trade amount and the delivered principal amount.",
            source_key="fannie_good_delivery",
            category="Trading & Settlement",
        ),
        make_row(
            "WAC",
            definition="Short for weighted-average coupon; the balance-weighted average coupon rate associated with the pooled loans or security disclosure field.",
            source_key="fannie_disclosure_guide",
            category="Rates, Cash Flow & Analytics",
        ),
        make_row(
            "WALA",
            definition="Short for weighted-average loan age; the balance-weighted average number of months since the loans were originated or modified.",
            source_key="fannie_disclosure_guide",
            category="Rates, Cash Flow & Analytics",
        ),
        make_row(
            "WAM",
            definition="Short for weighted-average maturity; a general market shorthand for the balance-weighted remaining maturity of the pooled collateral or security.",
            source_key="fannie_disclosure_guide",
            category="Rates, Cash Flow & Analytics",
        ),
        make_row(
            "Weighted-average coupon",
            definition="The balance-weighted average interest rate of the loans in the pool or of the related disclosed coupon metric.",
            source_key="fannie_disclosure_guide",
            category="Rates, Cash Flow & Analytics",
        ),
        make_row(
            "Weighted-average loan age",
            definition="The balance-weighted average number of months from origination or modification to the reporting date.",
            source_key="fannie_disclosure_guide",
            category="Rates, Cash Flow & Analytics",
        ),
        make_row(
            "Weighted-average maturity",
            definition="The balance-weighted average remaining term of the pooled mortgages, often used as a shorthand portfolio measure.",
            source_key="fannie_disclosure_guide",
            category="Rates, Cash Flow & Analytics",
        ),
        make_row(
            "Whole loan",
            definition="An individual mortgage loan sold as a loan asset rather than delivered inside an MBS pool.",
            source_key="fannie_glossary_w",
        ),
        make_row(
            "Yield",
            definition="A return measure used in pricing commitments and securities; in agency MBS operations, yield conventions help determine prices, pair-offs, and execution economics.",
            source_key="fannie_c3_1_01",
            category="Rates, Cash Flow & Analytics",
        ),
    ]
    return rows


def dedupe_rows(rows: Iterable[dict[str, str]]) -> list[dict[str, str]]:
    seen: set[str] = set()
    deduped: list[dict[str, str]] = []
    for row in rows:
        key = normalize_term(row["Term"])
        if key in seen:
            continue
        seen.add(key)
        deduped.append(row)
    return deduped


def build_workbook(rows: list[dict[str, str]]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Glossary"

    headers = ["Term", "Acronym", "Category", "Definition", "Notes", "Source Type", "Source Title", "Source URL"]
    ws.append(headers)
    for row in rows:
        ws.append([row[h] for h in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    wrap = Alignment(vertical="top", wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap

    widths = {
        "A": 34,
        "B": 12,
        "C": 30,
        "D": 90,
        "E": 48,
        "F": 22,
        "G": 44,
        "H": 78,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{ws.max_row}"

    table = Table(displayName="MBSGlossary", ref=f"A1:H{ws.max_row}")
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

    about = wb.create_sheet("About")
    about.append(["Field", "Value"])
    about.append(["Workbook", "MBS Pooling Dictionary"])
    about.append(["Scope", "Broad working glossary for single-family agency MBS pooling, centered on Fannie Mae terminology and the user's PDF."])
    about.append(["Generated File", OUTPUT_PATH.name])
    about.append(["Row Count", str(len(rows))])
    about.append(["Notes", "This is broad rather than literally universal. It is designed to be exhaustive for learning the pooling workflow, not every disclosure field in every structured product."])
    for cell in about[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap
    about.column_dimensions["A"].width = 24
    about.column_dimensions["B"].width = 110
    for row in about.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap

    sources_ws = wb.create_sheet("Sources")
    sources_ws.append(["Source Key", "Title", "URL", "Local Copy"])
    for key, source in sorted(SOURCES.items()):
        local_copy = str(source.local_path.relative_to(ROOT)) if source.local_path else ""
        sources_ws.append([key, source.title, source.url, local_copy])
    for cell in sources_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap
    for col, width in {"A": 24, "B": 54, "C": 96, "D": 34}.items():
        sources_ws.column_dimensions[col].width = width
    for row in sources_ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap

    for sheet in (ws, about, sources_ws):
        for idx in range(1, sheet.max_row + 1):
            sheet.row_dimensions[idx].height = 30 if idx == 1 else 42

    return wb


def main() -> None:
    rows = dedupe_rows(official_rows() + manual_rows())
    rows.sort(key=lambda row: normalize_term(row["Term"]))
    wb = build_workbook(rows)
    wb.save(OUTPUT_PATH)
    print(f"Wrote {len(rows)} glossary rows to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
